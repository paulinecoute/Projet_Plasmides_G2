import os
import openpyxl
from django.core.management.base import BaseCommand
from django.conf import settings
from django.core.files import File
from biolib.models import Plasmid, PlasmidCollection, User, CampaignTemplate, TemplatePart

class Command(BaseCommand):
    help = 'Charge les plasmides et parse intelligemment les templates Excel'

    def handle(self, *args, **kwargs):
        data_dir = os.path.join(settings.BASE_DIR, 'data_web')
        self.stdout.write("--- Démarrage de l'import Intelligent ---")

        if not os.path.exists(data_dir):
            self.stdout.write(self.style.ERROR(f"ERREUR : '{data_dir}' n'existe pas !"))
            return

        admin_user = User.objects.filter(is_superuser=True).first()
        if not admin_user:
            self.stdout.write(self.style.ERROR("Erreur : Aucun administrateur trouvé."))
            return

        count_plasmids = 0
        count_templates = 0

        for root, dirs, files in os.walk(data_dir):
            folder_name = os.path.basename(root)

            # 1. Gestion Collection
            collection = None
            if folder_name != 'data_web':
                collection_name = f"Collection {folder_name}"
                collection, created = PlasmidCollection.objects.get_or_create(
                    name=collection_name,
                    defaults={'owner': admin_user, 'publication_status': 'approved'}
                )

            # 2. Filtrage Fichiers
            valid_files = [f for f in files if f.lower().endswith(('.gb', '.dna', '.fasta', '.xlsx'))]
            
            for filename in valid_files:
                file_path = os.path.join(root, filename)
                identifier = os.path.splitext(filename)[0]

                # ==========================================================
                # TRAITEMENT TEMPLATE (.xlsx) -> ON LE PARSE !
                # ==========================================================
                if filename.lower().endswith('.xlsx'):
                    try:
                        template = CampaignTemplate.objects.filter(name=identifier, owner=admin_user).first()
                        
                        if not template:
                            template = CampaignTemplate(
                                name=identifier,
                                owner=admin_user,
                                description="",
                                visibility='public',
                                is_public=True
                            )
                        else:
                            template.description = "" # Nettoyage description

                        # Sauvegarde physique du fichier
                        with open(file_path, 'rb') as f_byte:
                            template.file.save(filename, File(f_byte), save=True)

                        # --- PARSING INTELLIGENT ---
                        wb = openpyxl.load_workbook(file_path, data_only=True)
                        ws = wb.active

                        # 1. Métadonnées (Enzyme/Séparateur - Positions fixes standards)
                        # B2 = Enzyme, B4 = Séparateur
                        enzyme_val = ws['B2'].value
                        separator_val = ws['B4'].value
                        if enzyme_val: template.enzyme = str(enzyme_val).strip()
                        if separator_val: template.output_separator = str(separator_val).strip()
                        template.save()

                        # 2. RECHERCHE DE L'ANCRE "Part name ->"
                        # On ne suppose pas la ligne, on la cherche.
                        start_row = 0
                        start_col = 0
                        found = False

                        # On scanne les 20 premières lignes et 5 premières colonnes
                        for r in range(1, 20):
                            for c in range(1, 6):
                                cell_val = str(ws.cell(row=r, column=c).value).strip()
                                if "Part name" in cell_val and "->" in cell_val:
                                    start_row = r
                                    start_col = c + 1 # Les données commencent la colonne d'après
                                    found = True
                                    break
                            if found: break
                        
                        if not found:
                            self.stdout.write(self.style.WARNING(f"   ! Impossible de trouver 'Part name ->' dans {filename}. Parsing annulé."))
                            continue

                        # 3. CRÉATION DES PARTIES
                        # On supprime les anciennes pour recréer proprement
                        template.parts.all().delete()

                        # Définition des lignes relatives à l'ancre trouvée
                        row_name = start_row
                        row_type = start_row + 1
                        row_opt  = start_row + 2
                        row_incl = start_row + 3

                        current_col = start_col
                        order_counter = 1

                        while True:
                            # Lecture du Nom
                            part_name = ws.cell(row=row_name, column=current_col).value

                            # Condition d'arrêt : Si vide ou si on tombe sur la section Output
                            if not part_name or str(part_name).strip() == "" or "output" in str(part_name).lower():
                                break

                            # Lecture Type
                            part_type = ws.cell(row=row_type, column=current_col).value or "1"
                            
                            # Lecture Optionnel (Excel : True = Optionnel)
                            val_opt = str(ws.cell(row=row_opt, column=current_col).value).lower()
                            # Django : is_mandatory (Inverse de optionnel)
                            is_mandatory = False if val_opt == 'true' else True

                            # Lecture Include Output
                            val_inc = str(ws.cell(row=row_incl, column=current_col).value).lower()
                            include_in_output = True if val_inc == 'true' else False

                            # Création en base
                            TemplatePart.objects.create(
                                template=template,
                                name=str(part_name),
                                type_id=str(part_type),
                                order=order_counter,
                                is_mandatory=is_mandatory,
                                include_in_output=include_in_output,
                                is_separable=False # Valeur par défaut
                            )

                            current_col += 1
                            order_counter += 1

                        count_templates += 1
                        self.stdout.write(f"   # Template OK : {filename} ({order_counter-1} parties générées)")

                    except Exception as e:
                         self.stdout.write(self.style.ERROR(f"Erreur Template {filename}: {e}"))
                    
                    continue

                # ==========================================================
                # TRAITEMENT PLASMIDE
                # ==========================================================
                try:
                    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f_read:
                        content_seq = f_read.read()

                    plasmid = Plasmid.objects.filter(identifier=identifier).first()
                    action = "Updated" if plasmid else "Created"

                    if not plasmid:
                        plasmid = Plasmid(identifier=identifier, name=identifier)
                    
                    plasmid.sequence = content_seq[:200] + "..."
                    
                    with open(file_path, 'rb') as f_byte:
                        plasmid.genbank_file.save(filename, File(f_byte), save=True)

                    if collection:
                        plasmid.collections.add(collection)

                    count_plasmids += 1
                    self.stdout.write(f"   + Plasmide {action} : {filename}")

                except Exception as e:
                    self.stdout.write(self.style.ERROR(f"Erreur Plasmide {filename}: {e}"))

        self.stdout.write(self.style.SUCCESS(f"--- TERMINE : {count_plasmids} plasmides, {count_templates} templates ---"))
