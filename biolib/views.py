from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.http import FileResponse, HttpResponse, Http404,HttpResponseForbidden
from django.conf import settings
from django.db.models import Q
from .forms import CustomUserCreationForm, SimulationForm, CampaignTemplateForm, TemplatePartFormSet, CorrespondenceForm, PlasmidForm, FeatureFormSet
from .models import Simulation, CampaignTemplate, Plasmid, Team, User, Correspondence, PlasmidCollection
import traceback
import pathlib
import glob
import os
import csv
import zipfile
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd
from django.core.files.base import ContentFile
import re
import shutil
from django.contrib import messages
from django.utils.safestring import mark_safe
from Bio import SeqIO
from io import StringIO
from django.urls import reverse
import io
from django.core.exceptions import PermissionDenied


import insillyclo.data_source
try:
    import insillyclo.observer
    BaseObserver = insillyclo.observer.InSillyCloObserver
except ImportError:
    class BaseObserver: pass
import insillyclo.simulator
try:
    from my_insillyclo.simulator import compute_all
except ImportError:
    def compute_all(*args, **kwargs): pass

class ConsoleObserver:
    def notify_message(self, message):
        print(f"[SIMULATION] {message}")
    def notify_progress(self, value):
        pass

class DjangoConsoleObserver(insillyclo.observer.InSillyCloCliObserver):
    def __init__(self):
        super().__init__(debug=False, fail_on_error=True)

    def notify_message(self, message):
        print(f"[INSILLYCLO] {message}")


# ==============================================================================
# 1. PAGES GÉNÉRALES
# ==============================================================================

def home(request):
    return render(request, 'biolib/home.html')

def search_view(request):
    """
    Moteur de recherche global (Templates, Simulations, Collections, Plasmides + Séquences)
    """
    query = request.GET.get('q', '')

    templates = CampaignTemplate.objects.none()
    simulations = Simulation.objects.none()
    plasmids = Plasmid.objects.none()
    collections = PlasmidCollection.objects.none()

    if query:
        # 1. DÉFINITION DES DROITS D'ACCÈS
        if request.user.is_authenticated:
            # Templates : Mes privés + Mon équipe + Publics
            tmpl_access = Q(owner=request.user) | Q(visibility='team', team__members=request.user) | Q(visibility='public')

            # Simulations : Mes privées + Mon équipe
            sim_access = Q(user=request.user) | Q(visibility='team', team__members=request.user)

            # Collections : Mes privées + Mon équipe + Publiques
            col_access = Q(owner=request.user) | Q(team__members=request.user) | Q(is_public=True)

        else:
            # Invité : Uniquement le contenu public
            tmpl_access = Q(visibility='public')
            sim_access = Q(pk__in=[])
            col_access = Q(is_public=True)

        # 2. EXÉCUTION DE LA RECHERCHE

        # Templates (Nom ou Description)
        templates = CampaignTemplate.objects.filter(tmpl_access).filter(
            Q(name__icontains=query) | Q(description__icontains=query)
        ).distinct()

        # Simulations (Nom) - Seulement si connecté pour l'instant
        if request.user.is_authenticated:
            simulations = Simulation.objects.filter(sim_access).filter(
                name__icontains=query
            ).distinct()

        # Collections (Nom)
        collections = PlasmidCollection.objects.filter(col_access).filter(
            name__icontains=query
        ).distinct()

        # Plasmides (Nom, ID, OU SÉQUENCE ADN)
        # On ne cherche que dans les collections accessibles
        accessible_col_ids = PlasmidCollection.objects.filter(col_access).values_list('id', flat=True)
        plasmids = Plasmid.objects.filter(collections__id__in=accessible_col_ids).filter(
            Q(name__icontains=query) |
            Q(identifier__icontains=query) |
            Q(sequence__icontains=query)
        ).distinct()[:20]

    return render(request, 'biolib/search_results.html', {
        'query': query,
        'templates': templates,
        'simulations': simulations,
        'collections': collections,
        'plasmids': plasmids,
    })

def signup(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('home')
    else:
        form = CustomUserCreationForm()
    return render(request, 'biolib/signup.html', {'form': form})

@login_required
def dashboard(request):
    collections_count = PlasmidCollection.objects.filter(owner=request.user).count()
    correspondences_count = Correspondence.objects.filter(owner=request.user).count()
    teams_count = request.user.teams.count()

    return render(request, "biolib/dashboard.html", {
        "collections_count": collections_count,
        "correspondences_count": correspondences_count,
        "teams_count": teams_count,
    })

# ==============================================================================
# 2. GESTION DES TEMPLATES
# ==============================================================================

def template(request):
    view_type = request.GET.get('view', 'recent')
    templates = CampaignTemplate.objects.none()
    title = "Templates récents"

    if request.user.is_authenticated:
        if view_type == 'private':
            templates = CampaignTemplate.objects.filter(owner=request.user, visibility='private').order_by('-created_at')
            title = "Mes templates privés"
        elif view_type == 'team':
            templates = CampaignTemplate.objects.filter(visibility='team', team__members=request.user).distinct().order_by('-created_at')
            title = "Templates d'équipe"
        elif view_type == 'public':
            templates = CampaignTemplate.objects.filter(visibility='public').order_by('-created_at')
            title = "Templates publics"
        else: # recent
            templates = CampaignTemplate.objects.filter(
                Q(owner=request.user) |
                Q(visibility='team', team__members=request.user) |
                Q(visibility='public')
            ).distinct().order_by('-id')[:5]
            title = "Templates récents"
    else:
        # Invité
        anon_ids = request.session.get('anon_templates', [])
        if view_type == 'public':
            templates = CampaignTemplate.objects.filter(visibility='public').order_by('-id')
            title = "Templates publics"
        else:
            templates = CampaignTemplate.objects.filter(
                Q(id__in=anon_ids) | Q(visibility='public')
            ).distinct().order_by('-id')[:5]
            title = "Templates récents"

    return render(request, 'biolib/template.html', {
        'templates': templates,
        'current_view': view_type,
        'page_title': title
    })

def create_template(request):
    # On définit proprement l'utilisateur à passer au formulaire
    # Si connecté -> request.user
    # Si invité -> None (pour éviter le crash "AnonymousUser")
    form_user = request.user if request.user.is_authenticated else None

    if request.method == 'POST':
        # On utilise form_user au lieu de request.user
        form = CampaignTemplateForm(request.POST, request.FILES, user=form_user)
        formset = TemplatePartFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            template = form.save(commit=False)

            if request.user.is_authenticated:
                template.owner = request.user
            else:
                template.owner = None
                template.visibility = 'private'

            # Sécurité : Si Team choisi mais pas d'équipe sélectionnée -> Force Privé
            if template.visibility == 'team' and not template.team:
                template.visibility = 'private'

            template.save()

            parts = formset.save(commit=False)
            for part in parts:
                part.template = template
                part.save()

            if not request.user.is_authenticated:
                session_templates = request.session.get('anon_templates', [])
                session_templates.append(template.id)
                request.session['anon_templates'] = session_templates
                request.session.modified = True

            return redirect('template')

    else:
        clone_id = request.GET.get('clone_from')

        if clone_id:
            original = get_object_or_404(CampaignTemplate, pk=clone_id)

            # Pré-remplissage (avec form_user)
            form = CampaignTemplateForm(user=form_user, initial={
                'name': f"{original.name} (Copie)",
                'enzyme': original.enzyme,
                'output_separator': original.output_separator,
                'description': original.description,
                'visibility': 'private', # Reset visibilité par sécurité
                'team': None             # Reset équipe
            })

            original_parts = original.parts.all().order_by('order')
            parts_data = []
            for part in original_parts:
                parts_data.append({
                    'name': part.name,
                    'type_id': part.type_id,
                    'order': part.order,
                    'is_mandatory': part.is_mandatory,
                    'include_in_output': part.include_in_output,
                    'is_separable': part.is_separable
                })

            formset = TemplatePartFormSet(initial=parts_data)
            formset.extra = len(parts_data)

        else:
            # Formulaire vide (avec form_user)
            form = CampaignTemplateForm(user=form_user)
            formset = TemplatePartFormSet()

    return render(request, 'biolib/create_template.html', {
        'form': form,
        'formset': formset
    })
def template_detail(request, pk):
    template = get_object_or_404(CampaignTemplate, pk=pk)
    return render(request, 'biolib/template_detail.html', {'template': template})

@login_required
def delete_template(request, pk):
    template = get_object_or_404(CampaignTemplate, pk=pk)

    if template.visibility == 'public':
        if not request.user.is_staff:
            return HttpResponse("Accès refusé.", status=403)
    elif template.visibility == 'team':
        is_team_leader = (template.team and template.team.leader == request.user)
        if request.user != template.owner and not is_team_leader:
             return HttpResponse("Accès refusé.", status=403)
    else:
        if request.user != template.owner:
            return HttpResponse("Accès refusé.", status=403)

    if request.method == 'POST':
        template.delete()
        return redirect('template')

    return render(request, 'biolib/template_confirm_delete.html', {'template': template})

def export_template_excel(request, template_id):
    template = get_object_or_404(CampaignTemplate, id=template_id)
    parts = template.parts.all().order_by('order')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assembly Template"

    blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    ws['A1'] = "Assembly settings"; ws['A1'].font = bold_font; ws['A1'].fill = blue_fill
    ws['A2'] = "Restriction enzyme"; ws['A2'].font = bold_font; ws['A2'].fill = blue_fill
    ws['B2'] = template.enzyme; ws['B2'].fill = green_fill
    ws['A3'] = "Name"; ws['A3'].font = bold_font; ws['A3'].fill = blue_fill
    ws['B3'] = template.name; ws['B3'].fill = green_fill
    ws['A4'] = "Output separator"; ws['A4'].font = bold_font; ws['A4'].fill = blue_fill
    ws['B4'] = template.output_separator; ws['B4'].fill = green_fill

    base_row = 8
    headers = ["Assembly composition", "Part types ->", "Is optional part ->", "Part name should be in output name ->", "Output plasmid id ↓"]
    for i, text in enumerate(headers):
        cell = ws.cell(row=base_row + i, column=1, value=text)
        cell.font = bold_font; cell.fill = blue_fill

    for index, part in enumerate(parts):
        col_num = 2 + index
        c1 = ws.cell(row=base_row, column=col_num, value=part.name); c1.fill = green_fill; c1.alignment = center_align
        c2 = ws.cell(row=base_row + 1, column=col_num, value=part.type_id); c2.fill = green_fill; c2.alignment = center_align
        c3 = ws.cell(row=base_row + 2, column=col_num, value="False" if part.is_mandatory else "True"); c3.fill = green_fill; c3.alignment = center_align
        c4 = ws.cell(row=base_row + 3, column=col_num, value="True" if part.include_in_output else "False"); c4.fill = green_fill; c4.alignment = center_align
        c5 = ws.cell(row=base_row + 4, column=col_num, value="↓"); c5.fill = blue_fill; c5.alignment = center_align

    ws.column_dimensions['A'].width = 35
    for col in range(2, 2 + len(parts)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    clean_name = template.name.replace(" ", "_")
    response['Content-Disposition'] = f'attachment; filename="Campaign_{clean_name}.xlsx"'
    wb.save(response)
    return response

#### TEMPLATES LIÉES AUX ÉQUIPES
@login_required
def team_templates(request, team_id):
    team = get_object_or_404(
        Team,
        id=team_id,
        members=request.user
    )

    templates = CampaignTemplate.objects.filter(
        visibility='team',
        team=team
    ).order_by('-created_at')

    return render(request, "biolib/template.html", {
        "templates": templates,
        "current_view": "team",
        "page_title": f"Templates de l’équipe {team.name}",
    })



def simulation_list(request):
    view_type = request.GET.get('view', 'recent')
    simulations = Simulation.objects.none()
    title = "Simulations récentes"

    if request.user.is_authenticated:
        # CAS CONNECTÉ : On prend tout depuis la BDD liée à l'user
        source_qs = Simulation.objects.filter(user=request.user)
        team_qs = Simulation.objects.filter(visibility='team', team__members=request.user)
    else:
        # CAS INVITÉ : On prend tout depuis la Session
        sim_ids = request.session.get('anonymous_simulations', [])
        source_qs = Simulation.objects.filter(id__in=sim_ids)
        team_qs = Simulation.objects.none() # Un invité n'a pas d'équipe

    if view_type == 'recent':
        # On affiche les 5 dernières (BDD ou Session)
        simulations = source_qs.order_by('-date_run')[:5]
        title = "Simulations récentes"

    elif view_type == 'team':
        # Uniquement si connecté, sinon vide (mais la page reste la même)
        simulations = team_qs.distinct().order_by('-date_run')
        title = "Simulations d'équipe"

    else: # view == 'mine'
        # Tout l'historique (BDD ou Session)
        simulations = source_qs.order_by('-date_run')
        title = "Mes simulations"

    return render(request, 'biolib/simulation_list.html', {
        'simulations': simulations,
        'current_view': view_type,
        'page_title': title
    })

def create_simulation(request):
    if request.method == 'POST':
        form = SimulationForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            simulation = form.save(commit=False)

            if request.user.is_authenticated:
                simulation.user = request.user
            else:
                simulation.user = None

            simulation.status = 'RUNNING'

            # Gestion des enzymes et primers
            selected_enzymes = form.cleaned_data.get('custom_enzymes')
            simulation.custom_enzymes = ",".join(selected_enzymes) if selected_enzymes else ""
            simulation.pcr_primers = form.cleaned_data.get('pcr_primers')

            # Gestion Visibilité
            if not simulation.visibility:
                simulation.visibility = 'private'
            if simulation.visibility == 'team' and not simulation.team:
                simulation.visibility = 'private'

            simulation.save()
            form.save_m2m() # Important pour les relations ManyToMany du formulaire

            # Gestion session anonyme
            if not request.user.is_authenticated:
                anon_sims = request.session.get('anonymous_simulations', [])
                anon_sims.append(simulation.id)
                request.session['anonymous_simulations'] = anon_sims
                request.session.modified = True

            # Création du dossier de travail
            output_folder = os.path.join(settings.MEDIA_ROOT, 'simulations', str(simulation.id))
            os.makedirs(output_folder, exist_ok=True)

            path_xlsx = simulation.template_file.path
            path_csv_list = [simulation.campaign_file.path] if simulation.campaign_file else []

            # Récupération des choix utilisateurs
            should_save = form.cleaned_data.get('save_to_library')
            col_name_input = form.cleaned_data.get('new_collection_name')
            camp_name_input = form.cleaned_data.get('campaign_save_name')
            temp_name_input = form.cleaned_data.get('template_save_name')

            # ==============================================================================
            # SAUVEGARDE MÉTADONNÉES (Campagne & Template)
            # ==============================================================================

            # 1. Correspondance (Campagne)
            if should_save and simulation.campaign_file and request.user.is_authenticated:
                final_camp_name = camp_name_input if camp_name_input else f"Campagne - {simulation.name}"
                Correspondence.objects.create(
                    name=final_camp_name,
                    file=simulation.campaign_file,
                    owner=request.user
                )

            # 2. Template (Modèle Excel)
            if should_save and simulation.template_file and request.user.is_authenticated:
                final_temp_name = temp_name_input if temp_name_input else f"Template - {simulation.name}"

                # On détermine la visibilité du template comme celle de la simulation
                tpl_visibility = simulation.visibility
                tpl_is_public = (tpl_visibility == 'public')

                CampaignTemplate.objects.create(
                    name=final_temp_name,
                    file=simulation.template_file,
                    owner=request.user,
                    enzyme=simulation.enzyme,
                    team=simulation.team,
                    visibility=tpl_visibility,
                    is_public=tpl_is_public,
                    description=f"Template sauvegardé depuis la simulation '{simulation.name}'"
                )


            raw_paths_list = []

            updated_plasmids = []
            new_plasmids = []

            target_collection = None
            if should_save and request.user.is_authenticated:
                final_col_name = col_name_input if col_name_input else f"Import Simu #{simulation.id}"
                target_collection = PlasmidCollection.objects.create(
                    name=final_col_name,
                    owner=request.user,
                    description=f"Import complet depuis la simulation {simulation.name}"
                )

            selected_collections = simulation.collections.all()
            for collection in selected_collections:
                for p in collection.plasmids.all():
                    if p.genbank_file:
                        try:
                            if os.path.exists(p.genbank_file.path):
                                raw_paths_list.append(p.genbank_file.path)

                                if target_collection:
                                    p.collections.add(target_collection)

                        except Exception:
                            pass

            if simulation.zip_file:
                try:
                    zip_path = simulation.zip_file.path
                    extract_path = os.path.join(output_folder, 'extracted_parts')
                    os.makedirs(extract_path, exist_ok=True)

                    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                        zip_ref.extractall(extract_path)

                    for root, dirs, files in os.walk(extract_path):
                        for file_name in files:
                            if file_name.lower().endswith(".gb") or file_name.lower().endswith(".gbk"):
                                full_path = os.path.join(root, file_name)

                                raw_paths_list.append(full_path)

                                if target_collection:
                                    try:
                                        existing_plasmid = Plasmid.objects.filter(
                                            collections__owner=request.user,
                                            name=file_name
                                        ).distinct().first()

                                        with open(full_path, 'rb') as f_io:
                                            file_content = ContentFile(f_io.read(), name=file_name)

                                        if existing_plasmid:
                                            # CAS A : IL EXISTE
                                            existing_plasmid.genbank_file.save(file_name, file_content, save=False)
                                            existing_plasmid.save()

                                            # Ajout à la nouvelle collection
                                            existing_plasmid.collections.add(target_collection)
                                            updated_plasmids.append(file_name)

                                        else:
                                            # CAS B : NOUVEAU
                                            new_p = Plasmid.objects.create(
                                                name=file_name,
                                                genbank_file=file_content
                                            )
                                            new_p.collections.add(target_collection)
                                            new_plasmids.append(file_name)

                                    except Exception as e_db:
                                        print(f"Erreur BDD {file_name}: {e_db}")

                except Exception as e_zip:
                    print(f"ERREUR ZIP: {e_zip}")

            if target_collection:
                if len(new_plasmids) > 0:
                    messages.success(request, f"✅ {len(new_plasmids)} nouveaux plasmides importés dans '{target_collection.name}'.")

                if len(updated_plasmids) > 0:
                    details = "<br>".join(updated_plasmids[:5])
                    if len(updated_plasmids) > 5: details += f"<br>... et {len(updated_plasmids)-5} autres."
                    msg_text = (
                        f"<strong>ℹ️ Mise à jour de {len(updated_plasmids)} plasmides existants :</strong><br>"
                        f"Ils ont été mis à jour et <strong>ajoutés</strong> à la collection '{target_collection.name}'.<br>"
                        f"<div class='mt-2 small text-muted border-start border-3 ps-2'>{details}</div>"
                    )
                    messages.info(request, mark_safe(msg_text))

                # Petit message bonus pour confirmer que les plasmides des collections sélectionnées sont inclus
                if not simulation.zip_file and len(raw_paths_list) > 0:
                     messages.success(request, f"Collection créée par fusion des {len(raw_paths_list)} plasmides sélectionnés.")


            staging_dir = os.path.join(output_folder, 'staging_plasmids')
            os.makedirs(staging_dir, exist_ok=True)

            final_gb_paths_for_simulation = []
            seen_names = set()

            def clean_filename(name):
                return re.sub(r'[^\w\-]', '', name.replace(" ", "_"))

            for original_path in raw_paths_list:
                try:
                    filename_brut = os.path.basename(original_path)
                    name_without_ext = os.path.splitext(filename_brut)[0]
                    safe_name = clean_filename(name_without_ext)[:60]

                    if safe_name in seen_names:
                        continue

                    seen_names.add(safe_name)

                    # Vérification rapide Genbank
                    try:
                         with open(original_path, "r") as f:
                            if not "LOCUS" in f.readline(): pass
                    except:
                        continue

                    new_filename = f"{safe_name}.gb"
                    new_full_path = os.path.join(staging_dir, new_filename)

                    shutil.copy(original_path, new_full_path)
                    final_gb_paths_for_simulation.append(new_full_path)

                except Exception as e:
                    print(f"Erreur lecture fichier {original_path}: {e}")

            print(f"DEBUG: {len(final_gb_paths_for_simulation)} fichiers prêts dans le Staging.")

            if len(final_gb_paths_for_simulation) == 0:
                print("ERREUR CRITIQUE: Aucun plasmide valide trouvé.")
                simulation.status = 'FAILED'
                simulation.save()
                messages.error(request, "Erreur : Aucun fichier GenBank trouvé. Vérifiez votre ZIP ou sélectionnez une collection.")
                return redirect('simulation_list')

            def_conc = form.cleaned_data.get('default_concentration') or 200.0
            path_conc = form.cleaned_data.get('concentration_file')
            if simulation.concentration_file:
                path_conc = simulation.concentration_file.path

            try:
                observer = DjangoConsoleObserver()
                compute_all(
                    observer=observer,
                    settings=None,
                    input_template_filled=path_xlsx,
                    input_parts_files=path_csv_list,
                    gb_plasmids=final_gb_paths_for_simulation,
                    output_dir=output_folder,
                    data_source="Django",
                    assembly_enzyme=simulation.enzyme,
                    gel_enzymes=selected_enzymes if selected_enzymes else [],
                    user_primers=simulation.pcr_primers,
                    default_mass_concentration=def_conc,
                    concentration_file=path_conc
                )

                tous_les_csv = glob.glob(os.path.join(output_folder, "*.csv"))
                for csv_path in tous_les_csv:
                    try:
                        df_temp = pd.read_csv(csv_path, sep=None, engine='python')
                        df_temp.to_csv(csv_path, sep=';', decimal=',', index=False)
                    except Exception:
                        pass

                creer_archive_resultats_seulement(dossier_source=output_folder, simulation_id=simulation.id)

                simulation.status = 'COMPLETED'

                path_png = os.path.join(output_folder, 'digestion.png')
                path_svg = os.path.join(output_folder, 'digestion.svg')
                if os.path.exists(path_png):
                    simulation.result_file = f"simulations/{simulation.id}/digestion.png"
                elif os.path.exists(path_svg):
                    simulation.result_file = f"simulations/{simulation.id}/digestion.svg"

                simulation.save()
                return redirect('simulation_result', pk=simulation.id)

            except Exception as e:
                print(f"Erreur simulation: {e}")
                traceback.print_exc()
                simulation.status = 'FAILED'
                simulation.save()
                messages.error(request, "Une erreur est survenue pendant le calcul de la simulation.")
                return redirect('simulation_list')

        else:
            print("ERREUR FORMULAIRE:", form.errors)
    else:
        form = SimulationForm(user=request.user)

    return render(request, 'biolib/create_simulation.html', {'form': form})

def simulation_result(request, pk=None):
    simulation = get_object_or_404(Simulation, pk=pk)

    # Sécurité : Si privé, vérifier user. Si session, vérifier session.
    if simulation.visibility == 'private' and simulation.user and simulation.user != request.user:
        return HttpResponse("Accès refusé", status=403)

    output_folder = os.path.join(settings.BASE_DIR, 'media', 'simulations', str(pk))
    csv_path = os.path.join(output_folder, 'dilutions.csv')

    if not os.path.exists(csv_path):
        found_csvs = list(pathlib.Path(output_folder).glob("*.csv"))
        valid_csvs = [f for f in found_csvs if "concentration" not in f.name.lower()]
        if valid_csvs:
            csv_path = str(valid_csvs[0])

    csv_data = []
    if os.path.exists(csv_path):
        try:
            with open(csv_path, 'r', encoding='utf-8') as f:
                content = f.read()
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(content[:1024])
                    reader = csv.reader(f, dialect)
                except:
                    reader = csv.reader(f, delimiter=',')
                csv_data = list(reader)
        except Exception:
            pass

    base_path = pathlib.Path(output_folder)
    all_gb_files = list(base_path.glob("*.gb"))

    generated_files = []

    db_csv_path = base_path / 'DB_produced_plasmid.csv'

    if db_csv_path.exists():
        produced_ids = set()
        try:
            with open(db_csv_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f, delimiter=';')
                next(reader, None)
                for row in reader:
                    if row:
                        produced_ids.add(row[0].strip())
        except Exception as e:
            print(f"Erreur lecture CSV: {e}")

        for f in all_gb_files:
            if f.stem in produced_ids:
                generated_files.append(f.name)
    my_collections = []
    if request.user.is_authenticated:
        my_collections = PlasmidCollection.objects.filter(owner=request.user)

    return render(request, 'biolib/simulation_result.html', {'simulation': simulation,'csv_data': csv_data,       'results': generated_files, 'my_collections': my_collections
    })

@login_required
def save_generated_plasmid(request, simulation_id):
    if request.method == "POST":
        simulation = get_object_or_404(Simulation, id=simulation_id)

        filename = request.POST.get('filename')
        collection_id = request.POST.get('collection_id')
        new_collection_name = request.POST.get('new_collection_name')

        target_collection = None

        if new_collection_name and new_collection_name.strip():
            target_collection = PlasmidCollection.objects.create(
                owner=request.user,
                name=new_collection_name.strip(),
                description="Collection créée depuis les résultats de simulation."
            )
            messages.success(request, f"Nouvelle collection '{target_collection.name}' créée avec succès.")

        elif collection_id:
            target_collection = get_object_or_404(PlasmidCollection, id=collection_id, owner=request.user)

        else:
            messages.error(request, "Vous devez choisir une collection existante ou en créer une nouvelle.")
            return redirect('simulation_result', pk=simulation_id)

        sim_root = os.path.join(settings.MEDIA_ROOT, 'simulations', str(simulation.id))
        path_in_outputs = os.path.join(sim_root, 'outputs', filename)
        path_at_root = os.path.join(sim_root, filename)

        source_file_path = None
        if os.path.exists(path_in_outputs):
            source_file_path = path_in_outputs
        elif os.path.exists(path_at_root):
            source_file_path = path_at_root

        if source_file_path:
            try:
                with open(source_file_path, 'rb') as f:
                    file_content = f.read()

                identifier = filename

                plasmid, created = Plasmid.objects.get_or_create(
                    identifier=identifier,
                    defaults={
                        'name': filename,
                        'sequence': ""
                    }
                )

                if created:
                    plasmid.genbank_file.save(filename, ContentFile(file_content), save=True)

                plasmid.collections.add(target_collection)

                messages.success(request, f"Le plasmide '{identifier}' a bien été ajouté à '{target_collection.name}'.")

            except Exception as e:
                messages.error(request, f"Erreur technique : {e}")
        else:
            messages.error(request, f"Fichier introuvable : {filename}")

    return redirect('simulation_result', pk=simulation_id)

def update_simulation_gel(request, pk):
    simulation = get_object_or_404(Simulation, pk=pk)

    if request.method == 'POST':
        # 1. Mise à jour des enzymes
        new_enzymes = request.POST.getlist('gel_enzymes')
        simulation.custom_enzymes = ",".join(new_enzymes) if new_enzymes else ""
        simulation.save()

        # 2. Définition des dossiers
        output_folder = os.path.join(settings.MEDIA_ROOT, 'simulations', str(simulation.id))
        path_xlsx = simulation.template_file.path

        # Le fichier de campagne est optionnel
        path_csv_list = [simulation.campaign_file.path] if simulation.campaign_file else []

        staging_dir = os.path.join(output_folder, 'staging_plasmids')

        gb_plasmids_paths = []
        if os.path.exists(staging_dir):
            # On récupère tous les .gb du dossier staging
            gb_plasmids_paths = glob.glob(os.path.join(staging_dir, "*.gb"))
        else:
            # Fallback de sécurité (si le staging a été supprimé)
            print("ERREUR: Dossier Staging introuvable. Tentative de récupération via la BDD.")
            # Ici on pourrait essayer de retrouver via simulation.collections,
            # mais le staging est censé être la source de vérité.
            pass

        print(f"DEBUG UPDATE GEL: {len(gb_plasmids_paths)} plasmides utilisés pour le recalcul.")

        try:
            observer = DjangoConsoleObserver()
            compute_all(
                observer=observer,
                settings=None,
                input_template_filled=path_xlsx,
                input_parts_files=path_csv_list,

                gb_plasmids=gb_plasmids_paths, # <--- La liste correcte et filtrée

                output_dir=output_folder,
                data_source="Django",
                assembly_enzyme=simulation.enzyme,
                gel_enzymes=new_enzymes, # Les nouvelles enzymes choisies
                user_primers=simulation.pcr_primers,
                default_mass_concentration=200 # Ou récupérer la valeur de simulation si stockée
            )

            # Post-traitement CSV (Conversion , vers ;)
            tous_les_csv = glob.glob(os.path.join(output_folder, "*.csv"))
            for csv_path in tous_les_csv:
                try:
                    df_temp = pd.read_csv(csv_path, sep=None, engine='python')
                    df_temp.to_csv(csv_path, sep=';', decimal=',', index=False)
                except Exception:
                    pass

            # Gestion de l'image de résultat
            path_png = os.path.join(output_folder, 'digestion.png')
            path_svg = os.path.join(output_folder, 'digestion.svg')

            # On force la mise à jour du chemin pour éviter le cache navigateur
            import time
            timestamp = int(time.time())

            if os.path.exists(path_png):
                   simulation.result_file = f"simulations/{simulation.id}/digestion.png?t={timestamp}"
            elif os.path.exists(path_svg):
                   simulation.result_file = f"simulations/{simulation.id}/digestion.svg?t={timestamp}"

            simulation.save()

        except Exception as e:
            print(f"Erreur Update Gel: {e}")
            traceback.print_exc()

    return redirect('simulation_result', pk=simulation.id)

@login_required
def team_simulations(request, team_id):
    team = get_object_or_404(
        Team,
        id=team_id,
        members=request.user
    )

    simulations = Simulation.objects.filter(
        visibility='team',
        team=team
    ).order_by('-date_run')

    return render(request, "biolib/simulation_list.html", {
        "simulations": simulations,
        "current_view": "team",
        "page_title": f"Simulations de l’équipe {team.name}",
    })


# ==============================================================================
# 4. FICHIERS ET DOWNLOADS
# ==============================================================================

def creer_archive_resultats_seulement(dossier_source, simulation_id, fichiers_a_exclure=None):
    if fichiers_a_exclure is None: fichiers_a_exclure = []
    nom_zip = f"simulation_{simulation_id}_archive.zip"
    chemin_zip = os.path.join(dossier_source, nom_zip)

    candidats = glob.glob(os.path.join(dossier_source, "*.gb")) + glob.glob(os.path.join(dossier_source, "*.csv")) + glob.glob(os.path.join(dossier_source, "*.png"))

    noms_exclus = set(os.path.basename(f) for f in fichiers_a_exclure)
    fichiers_finaux = [f for f in candidats if os.path.basename(f) not in noms_exclus]

    if not fichiers_finaux: return None

    try:
        with zipfile.ZipFile(chemin_zip, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for fichier in fichiers_finaux:
                zipf.write(fichier, arcname=os.path.basename(fichier))
        return chemin_zip
    except Exception:
        return None

def download_specific_file(request, pk, filename):
    if ".." in filename or "/" in filename: raise Http404
    file_path = os.path.join(settings.BASE_DIR, 'media','simulations', str(pk), filename)
    if os.path.exists(file_path):
        response = FileResponse(open(file_path, 'rb'), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    raise Http404

def download_simulation_zip(request, pk):
    # 1. Définition du chemin de base
    sim_dir = pathlib.Path(settings.MEDIA_ROOT) / 'simulations' / str(pk)

    if not sim_dir.exists():
        raise Http404("Dossier de simulation introuvable")

    # 2. Préparation de la liste des IDs à garder (Basé sur le CSV)
    produced_ids = set()
    db_csv_path = sim_dir / 'DB_produced_plasmid.csv'

    if db_csv_path.exists():
        try:
            with open(db_csv_path, 'r', encoding='utf-8') as f:
                # Utilisation du délimiteur ';' comme demandé
                reader = csv.reader(f, delimiter=';')
                next(reader, None) # Sauter le header
                for row in reader:
                    if row:
                        # On stocke l'ID (ex: pSA001)
                        produced_ids.add(row[0].strip())
        except Exception as e:
            print(f"Erreur lecture CSV: {e}")

    # 3. Création du ZIP en mémoire
    buffer = io.BytesIO()

    with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:

        # A. Ajouter les fichiers .gb FILTRÉS
        for f in sim_dir.glob("*.gb"):
            # Si le CSV existe, on filtre strictement
            if produced_ids:
                if f.stem in produced_ids:
                    zip_file.write(f, arcname=f.name)
            # Sinon (secours), on prend tout ce qui commence par 'p' (ex: pFinal...)
            # pour éviter les tADH1 (templates)
            else:
                if f.name.startswith("p"):
                    zip_file.write(f, arcname=f.name)

        # B. Ajouter les autres fichiers utiles (Images, CSV, etc.)
        # J'ai retiré '.json' de cette liste
        for f in sim_dir.glob("*"):
            if f.suffix in ['.csv', '.png', '.jpg', '.svg']:
                zip_file.write(f, arcname=f.name)

    # 4. Envoi du fichier
    buffer.seek(0)
    filename = f"simulation_{pk}_resultats.zip"
    return FileResponse(buffer, as_attachment=True, filename=filename)

def download_simulation_csv(request, pk):
    # Backward compatibility
    return download_specific_file(request, pk, 'dilutions.csv')

# ==============================================================================
# 5. GESTION DES ÉQUIPES ET COLLECTIONS
# ==============================================================================

@login_required
def team_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            team = Team.objects.create(name=name, description=request.POST.get('description', ''), purpose=request.POST.get('purpose'), leader=request.user)
            team.members.add(request.user)
            return redirect('teams')
    return render(request, 'biolib/team_create.html')

@login_required
def team_detail(request, team_id):
    team = get_object_or_404(Team, id=team_id, members=request.user)
    return render(request, 'biolib/team_detail.html', {
        'team': team,
        'is_leader': team.leader == request.user,
        'collections_count': team.plasmidcollection_set.count(),
        'tables_count': team.correspondence_set.count(),
        'campaigns_count': team.simulations.count(),
        'plasmids_count': Plasmid.objects.filter(collections__team=team).distinct().count(),
    })

@login_required
def team_manage_members(request, team_id):
    team = get_object_or_404(Team, id=team_id, members=request.user)
    if team.leader != request.user: return HttpResponse("Accès refusé", status=403)
    if request.method == 'POST':
        try:
            user = User.objects.get(email=request.POST.get('email'))
            team.members.add(user)
        except User.DoesNotExist: pass
    return render(request, 'biolib/team_manage_members.html', {'team': team})

@login_required
def team_change_leader(request, team_id, user_id):
    team = get_object_or_404(Team, id=team_id, members=request.user)
    if team.leader != request.user: return HttpResponse("Accès refusé", status=403)
    new_leader = get_object_or_404(User, id=user_id)
    if new_leader in team.members.all():
        team.leader = new_leader
        team.save()
    return redirect('team_detail', team_id=team.id)

@login_required
def team_remove_member(request, team_id, user_id):
    team = get_object_or_404(Team, id=team_id, members=request.user)
    if team.leader != request.user: return HttpResponse("Accès refusé", status=403)
    user = get_object_or_404(User, id=user_id)
    if user != team.leader:
        team.members.remove(user)
    return redirect('team_manage_members', team_id=team.id)

@login_required
def team_leave(request, team_id):
    team = get_object_or_404(Team, id=team_id, members=request.user)
    if request.user != team.leader:
        team.members.remove(request.user)
        return redirect('teams')
    return HttpResponse("Le chef ne peut pas quitter.", status=400)

@login_required
def team_delete(request, team_id):
    team = get_object_or_404(Team, id=team_id)
    if team.leader == request.user:
        if request.method == 'POST':
            team.delete()
            return redirect('teams')
        return render(request, 'biolib/team_confirm_delete.html', {'team': team})
    return HttpResponse("Accès refusé", status=403)

@login_required
def team_list(request):
    teams = Team.objects.filter(members=request.user)

    for team in teams:
        team.members_count = team.members.count()

        team.collections_count = team.plasmidcollection_set.count()

        team.plasmids_count = Plasmid.objects.filter(collections__team=team).distinct().count()

        team.simulations_count = team.simulations.count()

    context = {
        'teams': teams,
    }

    return render(request, 'biolib/teams.html', context)

# ============================================================
# COLLECTIONS UTILISATEUR (équipe)
# ============================================================

@login_required
def team_collections(request, team_id):
    team = get_object_or_404(Team, id=team_id, members=request.user)
    collections = PlasmidCollection.objects.filter(team=team)

    next_url = request.GET.get("next")

    forbidden = [
        request.path,
        "/plasmids/teams/",
    ]

    if next_url in forbidden:
        next_url = None

    if not next_url:
        next_url = reverse("team_detail", args=[team.id])

    return render(request, "biolib/team_collections.html", {
        "team": team,
        "collections": collections,
        "is_leader": team.leader == request.user,
        "next": next_url,
    })



@login_required
def team_collection_detail(request, team_id, collection_id):
    team = get_object_or_404(Team, id=team_id, members=request.user)
    collection = get_object_or_404(
        PlasmidCollection,
        id=collection_id,
        team=team
    )

    return render(request, "biolib/team_collection_detail.html", {
        "team": team,
        "collection": collection,
        "is_owner": collection.owner == request.user
    })


@login_required
def team_collection_create(request, team_id):
    team = get_object_or_404(Team, id=team_id, leader=request.user)

    if request.method == "POST":
        owner = get_object_or_404(
            User,
            id=request.POST.get("owner"),
            teams=team
        )

        PlasmidCollection.objects.create(
            name=request.POST["name"],
            description=request.POST.get("description", ""),
            owner=owner,
            team=team
        )

        return redirect("team_collections", team_id=team.id)

    return render(request, "biolib/team_collection_create.html", {
        "team": team,
        "members": team.members.all()
    })

@login_required
def choose_team_for_plasmids(request):
    teams = request.user.teams.all()
    next_url = request.GET.get("next")

    return render(
        request,
        "biolib/choose_team_for_plasmids.html",
        {
            "teams": teams,
            "next": next_url
        }
    )


# ============================================================
# COLLECTIONS UTILISATEUR (hors équipe)
# ============================================================

@login_required
def collections_view(request):
    collections = PlasmidCollection.objects.filter(
        owner=request.user,
        team__isnull=True
    )
    return render(request, "biolib/collections.html", {"collections": collections})



@login_required
def collection_create(request):
    if request.method == "POST":
        collection = PlasmidCollection.objects.create(
            name=request.POST["name"],
            description=request.POST.get("description", ""),
            owner=request.user
        )

        return redirect("plasmid_collection_detail", pk=collection.id)

    return render(request, "biolib/collection_create.html")


@login_required
def collection_detail(request, collection_id):
    collection = get_object_or_404(PlasmidCollection, id=collection_id)
    return render(
        request,
        "biolib/collection_detail.html",
        {
            "collection": collection,
            "is_owner": collection.owner == request.user
        }
    )


@login_required
def plasmid_upload(request, collection_id):
    collection = get_object_or_404(
        PlasmidCollection,
        id=collection_id,
        owner=request.user,
        publication_status='draft'
    )

    next_url = request.GET.get("next")

    if request.method == "POST":
        next_url = request.POST.get("next")

        for f in request.FILES.getlist("files"):
            seq_str = ""
            try:
                # On lit le contenu pour SeqIO
                content = f.read().decode('utf-8', errors='ignore')
                f.seek(0) # Important : on rembobine le fichier après lecture

                record = SeqIO.read(StringIO(content), "genbank")
                seq_str = str(record.seq).upper()

                f.seek(0)
            except Exception:
                f.seek(0)

            identifier = f.name

            existing_plasmid = Plasmid.objects.filter(identifier=identifier).first()

            if existing_plasmid:
                # MISE À JOUR (Update)
                existing_plasmid.genbank_file.save(f.name, f, save=False)
                if seq_str:
                    existing_plasmid.sequence = seq_str
                existing_plasmid.save()
                plasmid = existing_plasmid
            else:
                plasmid = Plasmid.objects.create(
                    identifier=identifier,
                    name=identifier,
                    genbank_file=f,
                    sequence=seq_str
                )

            plasmid.collections.add(collection)

        if next_url and next_url != "None":
            return redirect(next_url)

        return redirect("plasmid_collection_detail", pk=collection.id)

    return render(request, "biolib/plasmid_upload.html", {
        "collection": collection,
        "next": next_url
    })


@login_required
def plasmid_delete(request, plasmid_id):
    plasmid = get_object_or_404(
        Plasmid,
        id=plasmid_id,
        collections__owner=request.user
    )

    collection = plasmid.collections

    if request.method == "POST":
        next_url = request.POST.get("next")
        plasmid.delete()
        if next_url and next_url != "None":
            return redirect(next_url)

        return redirect("plasmid_collection_detail", pk=collection.id)

    return redirect("plasmid_collection_detail", pk=collection.id)

@login_required
def remove_plasmid_from_collection(request, collection_id, plasmid_id):
    # On récupère la collection
    collection = get_object_or_404(PlasmidCollection, id=collection_id, owner=request.user)

    plasmid = get_object_or_404(Plasmid, id=plasmid_id)

    # ON LE DÉTACHE (On coupe le lien sans supprimer l'objet)
    collection.plasmids.remove(plasmid)

    messages.success(request, f"Le plasmide '{plasmid.name}' a été retiré de la collection.")
    return redirect('plasmid_collection_detail', pk=collection.id)

@login_required
def collection_delete(request, collection_id):
    collection = get_object_or_404(
        PlasmidCollection,
        id=collection_id,
        owner=request.user
    )

    if request.method == "POST":
        next_url = request.POST.get("next")
        team = collection.team
        collection.delete()

        if next_url:
            return redirect(next_url)

        if team:
            return redirect("team_collections", team_id=team.id)
        return redirect("collections")

@login_required
def plasmid_collection_delete(request, pk):
    collection = get_object_or_404(
        PlasmidCollection,
        pk=pk,
        owner=request.user
    )

    if request.method == "POST":
        collection.delete()
        return redirect("plasmid_collection_list")

    return redirect("plasmid_collection_detail", pk=pk)


# ============================================================
# CORRESPONDENCES UTILISATEUR (hors équipe)
# ============================================================

@login_required
def correspondences_view(request):
    view_type = request.GET.get('view', 'my')

    if view_type == 'my':
        correspondences = Correspondence.objects.filter(
            owner=request.user,
            team__isnull=True
        ).order_by('-uploaded_at')
        title = "Mes tables de correspondance"

    else:
        # fallback sécurité
        correspondences = Correspondence.objects.none()
        title = "Tables de correspondance"

    return render(request, "biolib/correspondences.html", {
        "correspondences": correspondences,
        "current_view": view_type,
        "page_title": title,
    })



@login_required
def correspondence_upload(request):
    if request.method == "POST":
        Correspondence.objects.create(
            name=request.POST["name"],
            file=request.FILES["file"],
            owner=request.user
        )
        return redirect("correspondences")

    return render(request, "biolib/correspondence_upload.html")


@login_required
def correspondence_detail(request, pk):
    table = get_object_or_404(
        Correspondence,
        id=pk,
        owner=request.user
    )

    return render(request, "biolib/correspondence_detail.html", {
        "table": table,
        "is_owner": True
    })


@login_required
def correspondence_view_file(request, correspondence_id):
    table = get_object_or_404(
        Correspondence,
        id=correspondence_id,
        owner=request.user
    )

    try:
        with open(table.file.path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()
    except Exception:
        content = "Erreur lecture."

    return render(request, "biolib/correspondence_view_file.html", {
        "table": table,
        "content": content
    })


# ============================================================
# ACTIONS COMMUNES (UTILISATEUR + ÉQUIPE)
# ============================================================

@login_required
def correspondence_attach_file(request, correspondence_id):
    table = get_object_or_404(
        Correspondence,
        id=correspondence_id,
        team__isnull=True
    )

    if table.owner != request.user:
        return HttpResponse("Accès refusé", status=403)

    if request.method == "POST" and request.FILES.get("file"):
        table.file = request.FILES["file"]
        table.save()

    return redirect("correspondence_detail", correspondence_id=table.id)



@login_required
def correspondence_remove_file(request, correspondence_id):
    table = get_object_or_404(
        Correspondence,
        id=correspondence_id,
        team__isnull=True
    )

    if table.owner != request.user:
        return HttpResponse("Accès refusé", status=403)

    if request.method == "POST":
        table.file.delete(save=False)
        table.file = None
        table.save()

    return redirect("correspondence_detail", correspondence_id=table.id)



@login_required
def correspondence_delete(request, correspondence_id):
    table = get_object_or_404(
        Correspondence,
        id=correspondence_id,
        owner=request.user
    )

    if request.method == "POST":
        next_url = request.POST.get("next")
        team = table.team
        table.delete()

        if next_url:
            return redirect(next_url)

        if team:
            return redirect("team_correspondences", team_id=team.id)

        return redirect("correspondences")



# ============================================================
# CORRESPONDENCES D’ÉQUIPE
# ============================================================

@login_required
def team_correspondences(request, team_id):
    team = get_object_or_404(Team, id=team_id, members=request.user)

    correspondences = (
        Correspondence.objects
        .filter(team=team)
        .select_related("owner")
        .order_by("-uploaded_at")
    )

    next_url = request.GET.get("next")

    forbidden = [
        request.path,
        "/choose-team-for-correspondences/",
    ]

    if next_url in forbidden:
        next_url = None

    if not next_url:
        next_url = reverse("team_detail", args=[team.id])

    return render(request, "biolib/team_correspondences.html", {
        "team": team,
        "correspondences": correspondences,
        "is_leader": team.leader == request.user,
        "next": next_url,
    })



@login_required
def team_correspondence_create(request, team_id):
    team = get_object_or_404(Team, id=team_id, leader=request.user)

    if request.method == "POST":
        owner = get_object_or_404(
            User,
            id=request.POST["owner"],
            teams=team
        )

        Correspondence.objects.create(
            name=request.POST["name"],
            description=request.POST.get("description", ""),
            owner=owner,
            team=team
        )

        return redirect("team_correspondences", team_id=team.id)

    return render(request, "biolib/team_correspondence_create.html", {
        "team": team,
        "members": team.members.all()
    })


@login_required
def team_correspondence_detail(request, team_id, correspondence_id):
    team = get_object_or_404(Team, id=team_id, members=request.user)

    table = get_object_or_404(
        Correspondence,
        id=correspondence_id,
        team=team
    )

    return render(request, "biolib/team_correspondence_detail.html", {
        "team": team,
        "table": table,
        "is_owner": table.owner == request.user
    })



@login_required
def team_correspondence_view_file(request, team_id, correspondence_id):
    team = get_object_or_404(Team, id=team_id, members=request.user)

    table = get_object_or_404(
        Correspondence,
        id=correspondence_id,
        team=team
    )

    try:
        with open(table.file.path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()
    except Exception:
        content = "Erreur lecture."

    return render(request, "biolib/team_correspondence_view_file.html", {
        "team": team,
        "table": table,
        "content": content,
        "is_owner": table.owner == request.user
    })


@login_required
def team_correspondence_attach_file(request, team_id, correspondence_id):
    team = get_object_or_404(Team, id=team_id, members=request.user)

    table = get_object_or_404(
        Correspondence,
        id=correspondence_id,
        team=team
    )

    if table.owner != request.user:
        return HttpResponse("Accès refusé", status=403)

    if request.method == "POST" and request.FILES.get("file"):
        table.file = request.FILES["file"]
        table.save()

    return redirect(
        "team_correspondence_detail",
        team_id=team.id,
        correspondence_id=table.id
    )


@login_required
def team_correspondence_remove_file(request, team_id, correspondence_id):
    team = get_object_or_404(Team, id=team_id, members=request.user)

    table = get_object_or_404(
        Correspondence,
        id=correspondence_id,
        team=team
    )

    if table.owner != request.user:
        return HttpResponse("Accès refusé", status=403)

    if request.method == "POST":
        table.file.delete(save=False)
        table.file = None
        table.save()

    return redirect(
        "team_correspondence_detail",
        team_id=team.id,
        correspondence_id=table.id
    )


@login_required
def team_correspondence_delete(request, team_id, correspondence_id):
    team = get_object_or_404(Team, id=team_id, members=request.user)

    table = get_object_or_404(
        Correspondence,
        id=correspondence_id,
        team=team,
        owner=request.user
    )

    if request.method == "POST":
        table.delete()
        return redirect("team_correspondences", team_id=team.id)

@login_required
def choose_team_for_correspondences(request):
    teams = request.user.teams.all()
    next_url = request.GET.get("next")

    return render(
        request,
        "biolib/choose_team_for_correspondences.html",
        {
            "teams": teams,
            "next": next_url
        }
    )



# VISUALISATION DE PLASMIDES


def plasmid_collection_list(request):
    if not request.user.is_authenticated:
        public_collections = PlasmidCollection.objects.filter(
            publication_status='approved'
        ).order_by('-id')

        return render(request, 'biolib/plasmid_collection_list.html', {
            'public_collections': public_collections,
            'my_collections': [],
            'team_collections': [],
            'page_title': "Collections publiques"
        })


    my_collections = PlasmidCollection.objects.filter(
        owner=request.user
    ).distinct().order_by('-id')

    team_collections = PlasmidCollection.objects.filter(
        team__members=request.user
    ).distinct().order_by('-id')

    public_collections = PlasmidCollection.objects.filter(
        publication_status='approved'
    ).order_by('-id')

    return render(request, 'biolib/plasmid_collection_list.html', {
        'my_collections': my_collections,
        'team_collections': team_collections,
        'public_collections': public_collections,
        'page_title': "Mes Collections"
    })


def plasmid_collection_detail(request, pk):
    collection = get_object_or_404(PlasmidCollection, pk=pk)

    is_owner = (request.user == collection.owner)
    is_public = (collection.publication_status == 'approved')
    is_admin = request.user.is_staff

    if not is_owner and not is_public and not is_admin:
        return HttpResponseForbidden("Accès refusé : Cette collection est privée.")

    plasmids = collection.plasmids.all()

    context = {
        'collection': collection,
        'plasmids': plasmids,
        'is_owner': is_owner
    }

    return render(request, 'biolib/plasmid_collection_detail.html', context)

# visualisation plasmide

def plasmid_visualize(request, plasmid_id):
    plasmid = get_object_or_404(Plasmid, id=plasmid_id)

    genbank_content = ""
    if plasmid.genbank_file:
        try:
            with open(plasmid.genbank_file.path, 'r', encoding='utf-8') as f:
                genbank_content = f.read()
        except Exception as e:
            print(f"Erreur de lecture : {e}")
            genbank_content = plasmid.sequence
    else:
        genbank_content = plasmid.sequence

    if request.user.is_authenticated:
        can_edit = plasmid.collections.filter(owner=request.user).exists()
    else:
        can_edit = False

    return render(request, 'biolib/plasmid_visualize.html', {
        'plasmid': plasmid,
        'genbank_content': genbank_content,
        'can_edit': can_edit,
    })


@login_required
def request_publication(request, pk):
    collection = get_object_or_404(PlasmidCollection, pk=pk, owner=request.user)

    if collection.publication_status == 'draft' or collection.publication_status == 'rejected':
        collection.publication_status = 'pending'
        collection.save()
        messages.success(request, f"La demande de publication pour '{collection.name}' a été envoyée aux administrateurs.")

    return redirect('plasmid_collection_detail', pk=pk)

# Vue ADMIN
@staff_member_required
def admin_publication_list(request):
    # Les collections
    pending_collections = PlasmidCollection.objects.filter(
        publication_status='pending'
    ).order_by('-id')

    # Les tables
    pending_tables = Correspondence.objects.filter(
        publication_status='pending'
    ).order_by('-id')

    return render(request, 'biolib/admin_publication_list.html', {
        'pending_collections': pending_collections,
        'pending_tables': pending_tables
    })

@staff_member_required
def admin_approve_collection(request, pk):
    collection = get_object_or_404(PlasmidCollection, pk=pk)

    if request.method == 'POST':
        collection.publication_status = 'approved'
        collection.admin_feedback = ""
        collection.save()
        messages.success(request, f"La collection '{collection.name}' est maintenant publique !")

    return redirect('admin_publication_list')

@staff_member_required
def admin_reject_collection(request, pk):
    collection = get_object_or_404(PlasmidCollection, pk=pk)

    if request.method == 'POST':
        reason = request.POST.get('reason')
        if reason:
            collection.publication_status = 'rejected'
            collection.admin_feedback = reason
            collection.save()
            messages.warning(request, f"La collection '{collection.name}' a été refusée.")
        else:
            messages.error(request, "Vous devez fournir une raison pour le refus.")

    return redirect('admin_publication_list')

def correspondence_list(request):

    my_tables = Correspondence.objects.filter(
        owner=request.user
    ).order_by('-id')

    team_tables = Correspondence.objects.filter(
        team__members=request.user
    ).exclude(owner=request.user).distinct().order_by('-id')

    public_tables = Correspondence.objects.filter(
        publication_status='approved'
    ).order_by('-id')

    context = {
        'my_tables': my_tables,
        'team_tables': team_tables,
        'public_tables': public_tables,
    }
    return render(request, 'biolib/correspondence_list.html', context)

@login_required
def correspondence_request_publication(request, pk):
    table = get_object_or_404(Correspondence, pk=pk)

    if request.user != table.owner:
        return HttpResponseForbidden("Non autorisé")

    if table.publication_status in ['draft', 'rejected']:
        table.publication_status = 'pending'
        table.save()
        # Redirection vers la page de détail de la table
        return redirect('correspondence_detail', pk=pk)

    return redirect('correspondence_list')

@login_required
def plasmid_edit(request, pk):
    plasmid = get_object_or_404(Plasmid, pk=pk)

    is_owner = plasmid.collections.filter(owner=request.user).exists()
    if not is_owner:
        raise PermissionDenied("Permission refusée.")

    genbank_record = None
    features_initial_data = []

    if plasmid.genbank_file:
        try:
            with plasmid.genbank_file.open('r') as handle:
                genbank_record = SeqIO.read(handle, "genbank")

            for feature in genbank_record.features:
                if feature.type == 'source': continue

                quals = feature.qualifiers
                current_name = quals.get('label', quals.get('gene', quals.get('product', quals.get('note', ['Inconnu']))))[0]

                identifier = f"{feature.type}::{feature.location}"

                features_initial_data.append({
                    'feature_identifier': identifier,
                    'feature_type': feature.type,
                    'feature_location': str(feature.location),
                    'feature_name': current_name
                })
        except Exception as e:
            print(f"Erreur lecture GenBank: {e}")

    if request.method == 'POST':
        plasmid_form = PlasmidForm(request.POST, request.FILES, instance=plasmid)
        feature_formset = FeatureFormSet(request.POST)

        # On vérifie que TOUT est valide
        if plasmid_form.is_valid() and feature_formset.is_valid():
            plasmid_instance = plasmid_form.save(commit=False)

            if genbank_record and feature_formset.has_changed():
                updated = False
                for form in feature_formset:
                    if form.has_changed() and 'feature_name' in form.cleaned_data:
                        target_id = form.cleaned_data['feature_identifier']
                        new_name = form.cleaned_data['feature_name']

                        for feature in genbank_record.features:
                            current_id = f"{feature.type}::{feature.location}"
                            if current_id == target_id:
                                feature.qualifiers['label'] = [new_name]
                                updated = True
                                break
                if updated:
                    output_buffer = io.StringIO()
                    SeqIO.write(genbank_record, output_buffer, "genbank")

                    new_content = ContentFile(output_buffer.getvalue().encode('utf-8'))

                    plasmid_instance.genbank_file.save(plasmid_instance.genbank_file.name, new_content, save=False)
                    print("Fichier GenBank mis à jour avec les nouveaux noms d'annotations.")

            plasmid_instance.save()
            return redirect('plasmid_visualize', plasmid_id=plasmid.pk)

    else:
        plasmid_form = PlasmidForm(instance=plasmid)
        feature_formset = FeatureFormSet(initial=features_initial_data)

    return render(request, 'biolib/plasmid_edit.html', {
        'form': plasmid_form,
        'feature_formset': feature_formset,
        'plasmid': plasmid
    })

@login_required
def plasmid_detail(request, pk):
    plasmid = get_object_or_404(Plasmid, pk=pk)

    can_edit = plasmid.collections.filter(owner=request.user).exists()

    return render(request, 'biolib/plasmid_detail.html', {
        'plasmid': plasmid,
        'can_edit': can_edit,
    })
@staff_member_required
def admin_correspondence_review(request, pk):
    table = get_object_or_404(Correspondence, pk=pk)

    return render(request, 'biolib/admin_table_review.html', {
        'table': table
    })
@staff_member_required
def admin_approve_correspondence(request, pk):
    if request.method == 'POST':
        table = get_object_or_404(Correspondence, pk=pk)
        table.publication_status = 'approved'
        table.admin_feedback = ""
        table.save()
    return redirect('admin_publication_list')

@staff_member_required
def admin_reject_correspondence(request, pk):
    if request.method == 'POST':
        table = get_object_or_404(Correspondence, pk=pk)
        reason = request.POST.get('reason', '')
        table.publication_status = 'rejected'
        table.admin_feedback = reason
        table.save()
    return redirect('admin_publication_list')
def correspondence_create(request):
    if request.method == 'POST':
        form = CorrespondenceForm(request.POST, request.FILES)
        if form.is_valid():
            table = form.save(commit=False)
            table.owner = request.user
            table.save()
            return redirect('correspondence_list')
    else:
        form = CorrespondenceForm()

    return render(request, 'biolib/correspondence_form.html', {'form': form})
